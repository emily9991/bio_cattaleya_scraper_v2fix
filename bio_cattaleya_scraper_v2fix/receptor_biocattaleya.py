# -*- coding: utf-8 -*-
"""
Bio Cattaleya — receptor local. Estructura fija bajo C:\\taobao_scraper
Flujo: extensión (JSON) → POST /guardar_completo → carpetas + master Excel.
"""
from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd
import json
import os
import re
import shutil
import requests
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

BASE_DIR = r"C:\taobao_scraper"
MASTER_XLSX = os.path.join(BASE_DIR, "master_biocattaleya.xlsx")
EXPORTS_DIR = os.path.join(BASE_DIR, "exports")
PRODUCTOS_DIR = os.path.join(BASE_DIR, "productos")
BACKUP_DIR = os.path.join(BASE_DIR, "_backups_master")

TASA_CAMBIO = 7.25
MARGEN = 2.5

os.makedirs(BASE_DIR, exist_ok=True)
os.makedirs(EXPORTS_DIR, exist_ok=True)
os.makedirs(PRODUCTOS_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

NOTION_API = "https://api.notion.com/v1"
NOTION_VERSION = "2022-06-28"


def load_env_file():
    """Carga .env junto al script o en C:\\taobao_scraper\\.env"""
    for p in (
        os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"),
        os.path.join(BASE_DIR, ".env"),
    ):
        if not os.path.isfile(p):
            continue
        try:
            with open(p, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    k, _, v = line.partition("=")
                    k, v = k.strip(), v.strip().strip('"').strip("'")
                    if k and k not in os.environ:
                        os.environ[k] = v
        except OSError as e:
            print(f"   .env: {e}")
        break


load_env_file()

# Orden fijo — mismas columnas en master y producto.xlsx (Galería Notion = URLs separadas por comas)
MASTER_COLUMNS = [
    "Tienda",
    "Nombre (ZH)",
    "Nombre (EN)",
    "Descripcion",
    "Specs",
    "Parametros",
    "Variaciones",
    "Costo CNY",
    "Costo USD",
    "Precio Venta USD",
    "Calificacion",
    "Sitio",
    "URL",
    "Categoria",
    "Galeria URLs (Notion)",
    "Imagen portada",
    "Notion OK",
    "Notion Page URL",
    "Imagen1",
    "Imagen2",
    "Imagen3",
    "Imagen4",
    "Imagen5",
    "Imagen6",
    "Imagen7",
    "Imagen8",
    "Variantes URLs",
    "本店推荐 JSON",
    "Img Desc1",
    "Img Desc2",
    "Img Desc3",
    "Video URL",
    "Video archivo local",
    "Fecha importacion",
    "Carpeta local",
    "Tasa CNY/USD",
]


def _urls_alternativas_descarga(url):
    if not url or not isinstance(url, str):
        return []
    u = url.strip()
    if u.startswith("//"):
        u = "https:" + u
    if not u.startswith("http"):
        return []
    seen = set()
    out = []

    def add(x):
        if x and x not in seen:
            seen.add(x)
            out.append(x)

    add(u)
    base_no_q = u.split("?")[0]
    if base_no_q != u:
        add(base_no_q)
    fixed = re.sub(
        r"(\.(?:jpg|jpeg|png))(\.(?:jpg|jpeg|png))?_\s*\.webp$",
        r"\1",
        base_no_q,
        flags=re.I,
    )
    if fixed != base_no_q:
        add(fixed)
    m = re.match(r"^(.+?)(_\d+x\d+\.(?:jpg|jpeg|png|webp))$", base_no_q, re.I)
    if m:
        add(m.group(1) + ".jpg")
    return out


def obtener_tasa_cambio():
    try:
        r = requests.get("https://api.exchangerate-api.com/v4/latest/CNY", timeout=5)
        return float(r.json()["rates"]["USD"])
    except Exception:
        return TASA_CAMBIO


def traducir_zh_a_en(texto):
    if not texto or not isinstance(texto, str):
        return ""
    t = texto.strip()
    if len(t) < 2:
        return t
    if not any("\u4e00" <= c <= "\u9fff" for c in t):
        return t
    try:
        r = requests.get(
            "https://translate.googleapis.com/translate_a/single",
            params={"client": "gtx", "sl": "zh-CN", "tl": "en", "dt": "t", "q": t[:5000]},
            timeout=12,
        )
        r.raise_for_status()
        data = r.json()
        if data and data[0]:
            parts = []
            for block in data[0]:
                if block and block[0]:
                    parts.append(block[0])
            return "".join(parts).strip()
    except Exception as e:
        print(f"   Traducción: {e}")
    return ""


def descargar_media(url, carpeta, nombre_archivo, referer_url="", es_video=False):
    try:
        if not url or not isinstance(url, str):
            return False
        referer = referer_url or "https://www.taobao.com/"
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Referer": referer,
            "Accept": "*/*" if es_video else "image/webp,image/apng,image/*,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "sec-fetch-dest": "video" if es_video else "image",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "cross-site",
        }
        session = requests.Session()
        if es_video:
            u0 = url.strip()
            if u0.startswith("//"):
                u0 = "https:" + u0
            candidatos = [u0]
        else:
            candidatos = _urls_alternativas_descarga(url)
        if not candidatos:
            return False
        ultimo_status = None
        for try_url in candidatos:
            r = session.get(try_url, stream=True, timeout=45, headers=headers)
            ultimo_status = r.status_code
            if r.status_code != 200:
                continue
            ruta = os.path.join(carpeta, nombre_archivo)
            with open(ruta, "wb") as f:
                for chunk in r.iter_content(8192):
                    if chunk:
                        f.write(chunk)
            if os.path.getsize(ruta) > 800:
                return True
            os.remove(ruta)
        print(f"   HTTP {ultimo_status} — {nombre_archivo}")
        return False
    except Exception as e:
        print(f"   Error descarga {nombre_archivo}: {e}")
        return False


def slug_carpeta(nombre):
    raw = (nombre or "producto")[:72]
    s = re.sub(r'[<>:"/\\|?*]', "", raw).strip()
    s = re.sub(r"\s+", "_", s)
    if not s:
        s = "producto"
    return s[:80]


def ruta_carpeta_producto(nombre):
    """C:\\taobao_scraper\\productos\\<nombre> o <nombre>_2 si existe."""
    base = slug_carpeta(nombre)
    os.makedirs(PRODUCTOS_DIR, exist_ok=True)
    path = os.path.join(PRODUCTOS_DIR, base)
    if not os.path.exists(path):
        return path
    for i in range(2, 500):
        alt = os.path.join(PRODUCTOS_DIR, f"{base}_{i}")
        if not os.path.exists(alt):
            return alt
    return os.path.join(PRODUCTOS_DIR, f"{base}_{int(datetime.now().timestamp())}")


def texto_parametros(data):
    if data.get("parametros_texto"):
        return str(data["parametros_texto"])[:8000]
    pr = data.get("parametros")
    if isinstance(pr, list):
        return " | ".join(str(x) for x in pr if x)[:8000]
    return ""


def urls_galeria_notion_csv(data):
    """Una celda, comas, compatible con galería Notion al pegar."""
    raw = (data.get("imagenes_galeria_notion") or "").strip()
    if raw:
        return raw[:32000]
    imgs = list(data.get("imagenes") or [])
    desc = list(data.get("imagenes_descripcion") or [])
    ivar = list(data.get("imagenes_variantes") or [])
    out, seen = [], set()
    for u in imgs + desc + ivar:
        if not u or not isinstance(u, str):
            continue
        u = u.strip()
        if u in seen:
            continue
        seen.add(u)
        out.append(u)
    return ",".join(out[:200])[:32000]


def normalizar_database_id_notion(raw):
    s = (raw or "").replace("-", "").strip()
    if len(s) != 32:
        return (raw or "").strip()
    return f"{s[:8]}-{s[8:12]}-{s[12:16]}-{s[16:20]}-{s[20:]}"


def enviar_a_notion(datos):
    """
    Crea una página en la base de datos de Notion. Requiere integración con acceso al DB.
    Variables: NOTION_TOKEN, NOTION_DATABASE_ID
    Opcionales: NOTION_PROP_TITLE, NOTION_PROP_URL, NOTION_PROP_CNY, NOTION_PROP_USD, NOTION_PROP_CAT
    """
    token = os.environ.get("NOTION_TOKEN", "").strip()
    db_raw = os.environ.get("NOTION_DATABASE_ID", "").strip()
    if not token or not db_raw:
        return False, "Notion no configurado (.env)", None, None

    if os.environ.get("NOTION_DISABLE", "").strip() in ("1", "true", "yes"):
        return False, "Notion desactivado (NOTION_DISABLE)", None, None

    db_id = normalizar_database_id_notion(db_raw)
    p_title = os.getenv("NOTION_PROP_TITLE", "Nombre")
    p_url = os.getenv("NOTION_PROP_URL", "URL Taobao")
    p_cny = os.getenv("NOTION_PROP_CNY", "Precio CNY")
    p_usd = os.getenv("NOTION_PROP_USD", "Precio USD")
    p_cat = os.getenv("NOTION_PROP_CAT", "Categoría")

    nombre = (datos.get("nombre") or "Producto")[:2000]
    try:
        precio = float(datos.get("precio") or 0)
    except (TypeError, ValueError):
        precio = 0.0
    tasa = obtener_tasa_cambio()
    usd = round(precio / tasa, 2) if precio else None
    url_tb = (datos.get("url") or "").strip()
    cat = datos.get("categoria_notion") or "Skincare"
    if cat not in ("Skincare", "Maquillaje"):
        cat = "Skincare"

    imgs = datos.get("imagenes") or []
    cover_url = ""
    if imgs and isinstance(imgs[0], str) and imgs[0].startswith("http"):
        cover_url = imgs[0][:2000]

    headers = {
        "Authorization": f"Bearer {token}",
        "Notion-Version": NOTION_VERSION,
        "Content-Type": "application/json",
    }

    properties = {
        p_title: {"title": [{"text": {"content": nombre}}]},
    }
    if url_tb:
        properties[p_url] = {"url": url_tb}
    if p_cny and precio > 0:
        properties[p_cny] = {"number": precio}
    if p_usd and usd is not None:
        properties[p_usd] = {"number": usd}
    if p_cat:
        properties[p_cat] = {"select": {"name": cat}}

    body = {"parent": {"database_id": db_id}, "properties": properties}
    if cover_url:
        body["cover"] = {"type": "external", "external": {"url": cover_url}}

    try:
        r = requests.post(
            f"{NOTION_API}/pages",
            headers=headers,
            json=body,
            timeout=35,
        )
        if r.status_code not in (200, 201):
            err = r.text[:900]
            return False, f"Notion HTTP {r.status_code}: {err}", None, None
        j = r.json()
        page_id = j.get("id") or ""
        page_url = ""
        if page_id:
            pid = page_id.replace("-", "")
            page_url = f"https://www.notion.so/{pid}"
        return True, "Sincronizado con Notion", page_id, page_url
    except Exception as e:
        return False, str(e), None, None


def backup_maestro():
    if os.path.exists(MASTER_XLSX):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(BACKUP_DIR, f"master_backup_{ts}.xlsx")
        shutil.copy2(MASTER_XLSX, dest)
        olds = sorted([f for f in os.listdir(BACKUP_DIR) if f.endswith(".xlsx")], reverse=True)
        for o in olds[12:]:
            try:
                os.remove(os.path.join(BACKUP_DIR, o))
            except OSError:
                pass


def aplicar_formato_excel(path):
    try:
        wb = load_workbook(path)
        ws = wb.active
        borde = Border(
            bottom=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="EEEEEE"),
        )
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            cell.fill = PatternFill("solid", fgColor="2D1B69")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borde
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = borde
                cell.font = Font(size=9, name="Calibri")
        for col_idx in range(1, len(MASTER_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = min(42, 12)
        ws.freeze_panes = "A2"
        wb.save(path)
    except Exception as e:
        print(f"   Formato Excel: {e}")


def fila_ordenada_desde_json(
    data,
    tasa,
    url_producto,
    carpeta_rel,
    video_local,
    fecha,
    notion_ok=False,
    notion_page_url="",
    notion_mensaje="",
):
    try:
        precio_cny = float(data.get("precio") or 0)
    except (TypeError, ValueError):
        precio_cny = 0.0
    precio_usd = round(precio_cny / tasa, 2) if precio_cny else ""
    precio_venta = round(float(precio_usd) * MARGEN, 2) if precio_usd != "" else ""

    nombre_zh = data.get("nombre") or ""
    nombre_en = nombre_zh
    if nombre_zh and any("\u4e00" <= c <= "\u9fff" for c in nombre_zh):
        tr = traducir_zh_a_en(nombre_zh)
        nombre_en = tr if tr else "(traducir)"

    tienda = data.get("tienda") or ""
    tienda = re.sub(r"\d+\.\d+", "", tienda)
    tienda = re.sub(r"好评率\d+%|平均\d+小时发货|客服满意度\d+%|88VIP", "", tienda).strip()

    specs = data.get("specs") or []
    specs_s = " | ".join(str(s) for s in specs)[:4000] if isinstance(specs, list) else str(specs)[:4000]

    vari = data.get("variaciones") or []
    vari_s = " | ".join(str(s) for s in vari)[:2000] if isinstance(vari, list) else str(vari)[:2000]

    imgs = data.get("imagenes") or []
    desc = data.get("imagenes_descripcion") or []
    ivar = data.get("imagenes_variantes") or []
    rec = data.get("tienda_recomendados") or []

    cat = data.get("categoria_notion") or "Skincare"
    if cat not in ("Skincare", "Maquillaje"):
        cat = "Skincare"
    galeria_csv = urls_galeria_notion_csv(data)
    portada = imgs[0] if len(imgs) > 0 and imgs[0] else ""

    if "no configurado" in (notion_mensaje or "").lower():
        notion_txt = "N/A"
        notion_url_cell = ""
    elif notion_ok:
        notion_txt = "Sí"
        notion_url_cell = notion_page_url or ""
    else:
        notion_txt = "No"
        notion_url_cell = ""

    row = {c: "" for c in MASTER_COLUMNS}
    row.update(
        {
            "Tienda": tienda,
            "Nombre (ZH)": nombre_zh,
            "Nombre (EN)": nombre_en,
            "Descripcion": str(data.get("descripcion") or "")[:4000],
            "Specs": specs_s,
            "Parametros": texto_parametros(data)[:8000],
            "Variaciones": vari_s,
            "Costo CNY": precio_cny if precio_cny else "",
            "Costo USD": precio_usd,
            "Precio Venta USD": precio_venta,
            "Calificacion": str(data.get("calificaciones") or ""),
            "Sitio": str(data.get("sitio") or ""),
            "URL": url_producto,
            "Categoria": cat,
            "Galeria URLs (Notion)": galeria_csv,
            "Imagen portada": portada,
            "Notion OK": notion_txt,
            "Notion Page URL": notion_url_cell,
            "Imagen1": imgs[0] if len(imgs) > 0 else "",
            "Imagen2": imgs[1] if len(imgs) > 1 else "",
            "Imagen3": imgs[2] if len(imgs) > 2 else "",
            "Imagen4": imgs[3] if len(imgs) > 3 else "",
            "Imagen5": imgs[4] if len(imgs) > 4 else "",
            "Imagen6": imgs[5] if len(imgs) > 5 else "",
            "Imagen7": imgs[6] if len(imgs) > 6 else "",
            "Imagen8": imgs[7] if len(imgs) > 7 else "",
            "Variantes URLs": " | ".join(str(u) for u in ivar[:24])[:4000],
            "本店推荐 JSON": json.dumps(rec, ensure_ascii=False)[:4000],
            "Img Desc1": desc[0] if len(desc) > 0 else "",
            "Img Desc2": desc[1] if len(desc) > 1 else "",
            "Img Desc3": desc[2] if len(desc) > 2 else "",
            "Video URL": str(data.get("video") or ""),
            "Video archivo local": video_local or "",
            "Fecha importacion": fecha,
            "Carpeta local": carpeta_rel,
            "Tasa CNY/USD": round(tasa, 4),
        }
    )
    return row


def dataframe_alineado(df):
    for c in MASTER_COLUMNS:
        if c not in df.columns:
            df[c] = ""
    return df[MASTER_COLUMNS]


def es_duplicado(df, url_producto):
    if df is None or df.empty or "URL" not in df.columns:
        return False
    u = str(url_producto or "").strip()
    return u in df["URL"].astype(str).values


@app.route("/guardar_completo", methods=["POST", "OPTIONS"])
def guardar_completo():
    if request.method == "OPTIONS":
        return jsonify({"status": "ok"})

    data = request.json or {}
    url_producto = (data.get("url") or "").strip()

    print("\n" + "=" * 60)
    print("  BIO CATTALEYA — guardar_completo (local)")
    print(f"  URL: {url_producto[:80]}")
    print("=" * 60)

    df_exist = None
    if os.path.exists(MASTER_XLSX):
        try:
            df_exist = pd.read_excel(MASTER_XLSX, engine="openpyxl")
        except PermissionError:
            msg = "El master_biocattaleya.xlsx está abierto. Ciérralo e inténtalo de nuevo."
            print(f"   {msg}")
            return jsonify({"status": "error", "message": msg}), 409
        except Exception as e:
            print(f"   Lectura master: {e}")

    if es_duplicado(df_exist, url_producto):
        return jsonify(
            {
                "status": "duplicate",
                "message": "Este producto (misma URL) ya está en el maestro.",
            }
        )

    tasa = obtener_tasa_cambio()
    nombre_prod = data.get("nombre") or "producto"
    product_root = ruta_carpeta_producto(nombre_prod)
    os.makedirs(product_root, exist_ok=True)
    fotos_dir = os.path.join(product_root, "fotos")
    os.makedirs(fotos_dir, exist_ok=True)

    carpeta_rel = os.path.relpath(product_root, BASE_DIR)

    ts_export = datetime.now().strftime("%Y%m%d_%H%M%S")
    export_snap = os.path.join(EXPORTS_DIR, f"export_{ts_export}.json")
    try:
        with open(export_snap, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"   exports JSON: {e}")

    json_local = os.path.join(product_root, "producto.json")
    with open(json_local, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    referer = url_producto or "https://www.taobao.com/"
    imgs = data.get("imagenes") or []
    n_ok = 0
    for i, u in enumerate(imgs, 1):
        if descargar_media(u, fotos_dir, f"galeria_{i:02d}.jpg", referer_url=referer):
            n_ok += 1

    desc_urls = data.get("imagenes_descripcion") or []
    d_ok = 0
    for i, u in enumerate(desc_urls, 1):
        if descargar_media(u, fotos_dir, f"detalle_{i:02d}.jpg", referer_url=referer):
            d_ok += 1

    ivar = data.get("imagenes_variantes") or []
    v_ok = 0
    for i, u in enumerate(ivar, 1):
        if descargar_media(u, fotos_dir, f"variante_{i:02d}.jpg", referer_url=referer):
            v_ok += 1

    video_url = (data.get("video") or "").strip()
    video_local = ""
    if video_url:
        if descargar_media(video_url, fotos_dir, "video.mp4", referer_url=referer, es_video=True):
            video_local = os.path.join("fotos", "video.mp4").replace("/", os.sep)

    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    notion_ok, notion_msg, notion_page_id, notion_page_url = enviar_a_notion(data)
    if notion_ok:
        print(f"   Notion ✓ {notion_page_url or notion_page_id}")
    else:
        print(f"   Notion: {notion_msg}")

    row = fila_ordenada_desde_json(
        data,
        tasa,
        url_producto,
        carpeta_rel,
        video_local,
        fecha,
        notion_ok=notion_ok,
        notion_page_url=notion_page_url or "",
        notion_mensaje=notion_msg or "",
    )

    df_new = pd.DataFrame([row], columns=MASTER_COLUMNS)
    backup_maestro()
    if df_exist is not None and not df_exist.empty:
        df_exist = dataframe_alineado(df_exist)
        df_final = pd.concat([df_exist, df_new], ignore_index=True)
    else:
        df_final = df_new

    try:
        df_final.to_excel(MASTER_XLSX, index=False, engine="openpyxl")
    except PermissionError:
        msg = "No se pudo escribir master_biocattaleya.xlsx (¿abierto en Excel?)."
        return jsonify({"status": "error", "message": msg}), 409

    aplicar_formato_excel(MASTER_XLSX)

    xlsx_producto = os.path.join(product_root, "producto.xlsx")
    try:
        df_new.to_excel(xlsx_producto, index=False, engine="openpyxl")
    except Exception as e:
        print(f"   producto.xlsx: {e}")

    print(f"   Carpeta: {carpeta_rel}")
    print(f"   Fotos galería OK: {n_ok}/{len(imgs)} · detalle: {d_ok}/{len(desc_urls)} · var: {v_ok}/{len(ivar)}")
    print("=" * 60)

    return jsonify(
        {
            "status": "success",
            "message": (
                f"Local OK · {carpeta_rel} · "
                f"Master {len(df_final)} filas · "
                f"{'Notion OK' if notion_ok else 'Notion: ' + (notion_msg or '')}"
            ),
            "carpeta": carpeta_rel,
            "total_maestro": len(df_final),
            "local_ok": True,
            "notion_ok": notion_ok,
            "notion_message": notion_msg,
            "notion_page_url": notion_page_url or "",
            "notion_page_id": notion_page_id or "",
        }
    )


@app.route("/stats", methods=["GET"])
def stats():
    if not os.path.exists(MASTER_XLSX):
        return jsonify({"total": 0, "mensaje": "master_biocattaleya.xlsx aún no existe"})
    try:
        df = pd.read_excel(MASTER_XLSX, engine="openpyxl")
        return jsonify({"total": len(df), "columnas": list(df.columns)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/", methods=["GET"])
def health():
    return jsonify({"status": "ok", "service": "BioCattaleya receptor", "base": BASE_DIR})


if __name__ == "__main__":
    print("=" * 60)
    print("  receptor_biocattaleya.py — local + Notion")
    print(f"  Raíz:     {BASE_DIR}")
    print(f"  Master:   {MASTER_XLSX}")
    print(f"  Exports:  {EXPORTS_DIR}")
    print(f"  Productos:{PRODUCTOS_DIR}")
    print("  POST /guardar_completo  → disco + Notion (si .env)")
    print("  NOTION_TOKEN / NOTION_DATABASE_ID en .env")
    print("=" * 60)
    app.run(host="127.0.0.1", port=5000, debug=False)
