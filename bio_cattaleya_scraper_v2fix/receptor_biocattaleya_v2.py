from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd
import json
import os
import re
import shutil
import requests
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

BASE_DIR    = r"C:\taobao_scraper"
OUTPUT_DIR  = os.path.join(BASE_DIR, "output")
EXCEL_PATH  = os.path.join(OUTPUT_DIR, "Maestro_BioCattaleya.xlsx")
BACKUP_DIR  = os.path.join(OUTPUT_DIR, "_backups")
TASA_CAMBIO = 7.25
MARGEN      = 2.5

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)


# ─── TASA DE CAMBIO ──────────────────────────────────────────
def obtener_tasa_cambio():
    try:
        r = requests.get("https://api.exchangerate-api.com/v4/latest/CNY", timeout=5)
        tasa = r.json()["rates"]["USD"]
        print(f"   Tasa CNY→USD: {tasa:.4f} (en línea)")
        return tasa
    except Exception:
        print(f"   Tasa CNY→USD: {TASA_CAMBIO} (fallback fijo)")
        return TASA_CAMBIO


# ─── VARIANTES DE URL (CDN Alibaba suele 404 si se normaliza mal) ──
def _urls_alternativas_descarga(url):
    """Genera candidatos: original completa, sin query, y nombre sin sufijo miniatura."""
    if not url or not isinstance(url, str):
        return []
    u = url.strip()
    if u.startswith("//"):
        u = "https:" + u
    if not u.startswith("http"):
        return []
    seen = set()
    out = []

    def add(x):
        if x and x not in seen:
            seen.add(x)
            out.append(x)

    add(u)
    base_no_q = u.split("?")[0]
    if base_no_q != u:
        add(base_no_q)
    # Doble extensión .jpg.jpg_.webp → .jpg (común en alicdn)
    fixed = re.sub(
        r"(\.(?:jpg|jpeg|png))(\.(?:jpg|jpeg|png))?_\s*\.webp$",
        r"\1",
        base_no_q,
        flags=re.I,
    )
    if fixed != base_no_q:
        add(fixed)
    # Sufijo _WxH.jpg al final (miniatura) → probar base + .jpg
    m = re.match(r"^(.+?)(_\d+x\d+\.(?:jpg|jpeg|png|webp))$", base_no_q, re.I)
    if m:
        add(m.group(1) + ".jpg")
    return out


# ─── TRADUCCIÓN ZH→EN (mismo endpoint que la extensión) ───────
def traducir_zh_a_en(texto):
    if not texto or not isinstance(texto, str):
        return ""
    t = texto.strip()
    if len(t) < 2:
        return t
    if not any("\u4e00" <= c <= "\u9fff" for c in t):
        return t
    try:
        r = requests.get(
            "https://translate.googleapis.com/translate_a/single",
            params={"client": "gtx", "sl": "zh-CN", "tl": "en", "dt": "t", "q": t[:5000]},
            timeout=12,
        )
        r.raise_for_status()
        data = r.json()
        if data and data[0]:
            parts = []
            for block in data[0]:
                if block and block[0]:
                    parts.append(block[0])
            return "".join(parts).strip()
    except Exception as e:
        print(f"   ⚠ Traducción falló: {e}")
    return ""


# ─── DESCARGA DE MEDIA CON ANTI-HOTLINK ──────────────────────
def descargar_media(url, carpeta, nombre_archivo, referer_url="", es_video=False):
    """
    Descarga una imagen o video respetando los headers anti-hotlink
    de Taobao/Tmall. Prueba varias variantes de URL si hay 404.
    Retorna True si tuvo éxito.
    """
    try:
        if not url or not isinstance(url, str):
            return False
        referer = referer_url or "https://www.taobao.com/"
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Referer": referer,
            "Accept": (
                "*/*"
                if es_video
                else "image/webp,image/apng,image/*,*/*;q=0.8"
            ),
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Accept-Encoding": "gzip, deflate, br",
            "sec-ch-ua": '"Chromium";v="124", "Google Chrome";v="124"',
            "sec-fetch-dest": "video" if es_video else "image",
            "sec-fetch-mode": "no-cors",
            "sec-fetch-site": "cross-site",
        }

        session = requests.Session()
        if es_video:
            u0 = url.strip()
            if u0.startswith("//"):
                u0 = "https:" + u0
            candidatos = [u0]
        else:
            candidatos = _urls_alternativas_descarga(url)
        if not candidatos:
            return False
        ultimo_status = None

        for try_url in candidatos:
            r = session.get(try_url, stream=True, timeout=30, headers=headers)
            ultimo_status = r.status_code
            if r.status_code != 200:
                continue
            ruta = os.path.join(carpeta, nombre_archivo)
            with open(ruta, "wb") as f:
                for chunk in r.iter_content(8192):
                    if chunk:
                        f.write(chunk)
            if os.path.getsize(ruta) > 1000:
                if try_url != candidatos[0]:
                    print(f"   ℹ {nombre_archivo}: OK con URL alternativa")
                return True
            os.remove(ruta)

        print(f"   ⚠ HTTP {ultimo_status} para {nombre_archivo}")
        return False

    except Exception as e:
        print(f"   ✗ Error descargando {nombre_archivo}: {e}")
        return False


# ─── BACKUP DEL EXCEL ────────────────────────────────────────
def hacer_backup_excel():
    if os.path.exists(EXCEL_PATH):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        destino = os.path.join(BACKUP_DIR, f"Maestro_backup_{ts}.xlsx")
        shutil.copy2(EXCEL_PATH, destino)
        backups = sorted(
            [f for f in os.listdir(BACKUP_DIR) if f.endswith(".xlsx")],
            reverse=True
        )
        for viejo in backups[10:]:
            os.remove(os.path.join(BACKUP_DIR, viejo))
        print(f"   Backup: {os.path.basename(destino)}")


# ─── DETECTAR DUPLICADO ──────────────────────────────────────
def es_duplicado(df, url_producto):
    if df is None or df.empty or "URL" not in df.columns:
        return False
    return url_producto in df["URL"].values


# ─── FORMATEAR EXCEL ─────────────────────────────────────────
def aplicar_formato_excel(path):
    try:
        wb = load_workbook(path)
        ws = wb.active

        COLOR_HEADER_BG  = "2D1B69"
        COLOR_HEADER_FG  = "FFFFFF"
        COLOR_FILA_PAR   = "F3EEFF"
        COLOR_FILA_IMPAR = "FFFFFF"

        borde = Border(
            bottom=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="EEEEEE")
        )

        for cell in ws[1]:
            cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=10, name="Calibri")
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borde
        ws.row_dimensions[1].height = 30

        col_nombres = [cell.value for cell in ws[1]]
        cols_url = [
            i + 1 for i, n in enumerate(col_nombres)
            if n and ("URL" in str(n) or "Imagen" in str(n) or "Img Desc" in str(n) or "Video" in str(n))
        ]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            color_fila = COLOR_FILA_PAR if row_idx % 2 == 0 else COLOR_FILA_IMPAR
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=color_fila)
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                cell.border = borde
                cell.font = Font(size=9, name="Calibri")
                if cell.column in cols_url and cell.value and str(cell.value).startswith("http"):
                    url_val = str(cell.value)
                    cell.hyperlink = url_val
                    cell.font = Font(size=9, name="Calibri", color="1155CC", underline="single")
                    cell.value = (url_val[:55] + "...") if len(url_val) > 55 else url_val
            ws.row_dimensions[row_idx].height = 18

        ANCHOS = {
            "Tienda": 22, "Nombre (ZH)": 35, "Nombre (EN)": 35,
            "Descripcion": 40, "Specs": 30, "Variaciones": 30,
            "Costo CNY": 12, "Costo USD": 12, "Precio Venta USD": 16,
            "Calificacion": 14, "Sitio": 16, "URL": 30,
            "Variantes URLs": 28, "本店推荐 JSON": 36,
        }
        for col_idx, col_cells in enumerate(ws.columns, 1):
            encabezado = str(col_cells[0].value or "")
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ancho = min(max(max_len + 2, 10), 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(ancho, ANCHOS.get(encabezado, 10))

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(path)
        print("   Formato Excel aplicado ✓")
    except Exception as e:
        print(f"   Advertencia al formatear Excel: {e}")


# ─── RUTA PRINCIPAL ──────────────────────────────────────────
@app.route('/guardar_completo', methods=['POST', 'OPTIONS'])
def guardar_completo():
    if request.method == 'OPTIONS':
        return jsonify({"status": "ok"})

    data = request.json
    url_producto = data.get("url", "")

    print("\n" + "=" * 60)
    print("  BIO CATTALEYA — Producto recibido")
    print(f"  URL: {url_producto[:70]}")
    print("=" * 60)

    # Leer Excel existente
    df_exist = None
    if os.path.exists(EXCEL_PATH):
        try:
            df_exist = pd.read_excel(EXCEL_PATH, engine="openpyxl")
        except PermissionError:
            msg = "❌ El Excel maestro está abierto en Windows. Ciérralo y reintenta."
            print(f"   {msg}")
            return jsonify({"status": "error", "message": msg}), 409
        except Exception as e:
            print(f"   Error al leer Excel: {e}")

    # Duplicado
    if es_duplicado(df_exist, url_producto):
        print(f"   ⚠ Duplicado detectado — no se agregó")
        return jsonify({
            "status": "duplicate",
            "message": "⚠️ Este producto ya existe en el Excel maestro (misma URL)."
        })

    tasa = obtener_tasa_cambio()

    # Crear carpeta del producto
    # Estructura: output / 20260401_223000_nombre_producto /
    #                         imagenes /
    #                         videos /
    #                         datos_completo.json
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_raw  = (data.get("nombre") or "sin_nombre")[:40]
    nombre_slug = re.sub(r'[^\w\s-]', '', nombre_raw).strip().replace(' ', '_')[:30]
    folder_name = f"{timestamp}_{nombre_slug}"
    product_path = os.path.join(OUTPUT_DIR, folder_name)

    img_folder   = os.path.join(product_path, "imagenes")
    video_folder = os.path.join(product_path, "videos")
    desc_folder  = os.path.join(product_path, "imagenes_descripcion")
    var_folder   = os.path.join(product_path, "imagenes_variantes")

    os.makedirs(img_folder,   exist_ok=True)
    os.makedirs(video_folder, exist_ok=True)
    os.makedirs(desc_folder,  exist_ok=True)
    os.makedirs(var_folder,   exist_ok=True)

    print(f"   Carpeta: {folder_name}")

    # Guardar JSON
    json_path = os.path.join(product_path, "datos_completo.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    print("   JSON guardado ✓")

    # Descargar imágenes del producto
    imagenes_urls = data.get("imagenes", [])
    imagenes_ok   = 0
    print(f"   Descargando {len(imagenes_urls)} imágenes del producto...")
    for i, url in enumerate(imagenes_urls, 1):
        ok = descargar_media(url, img_folder, f"foto_{i:02d}.jpg", referer_url=url_producto)
        if ok:
            imagenes_ok += 1
            print(f"     ✓ foto_{i:02d}.jpg")
        else:
            print(f"     ✗ foto_{i:02d} — falló")
    print(f"   Imágenes producto: {imagenes_ok}/{len(imagenes_urls)} ✓")

    # Imágenes de variantes (SKU / colores)
    imagenes_variantes = data.get("imagenes_variantes") or []
    var_ok = 0
    if imagenes_variantes:
        print(f"   Descargando {len(imagenes_variantes)} imágenes de variantes...")
        for i, vurl in enumerate(imagenes_variantes, 1):
            if descargar_media(vurl, var_folder, f"var_{i:02d}.jpg", referer_url=url_producto):
                var_ok += 1
        print(f"   Variantes: {var_ok}/{len(imagenes_variantes)} ✓")

    # Descargar imágenes de descripción
    desc_urls = data.get("imagenes_descripcion", [])
    desc_ok   = 0
    if desc_urls:
        print(f"   Descargando {len(desc_urls)} imágenes de descripción...")
        for i, url in enumerate(desc_urls, 1):
            ok = descargar_media(url, desc_folder, f"desc_{i:02d}.jpg", referer_url=url_producto)
            if ok:
                desc_ok += 1
        print(f"   Imágenes descripción: {desc_ok}/{len(desc_urls)} ✓")

    # Descargar video
    video_ok = False
    video_url = data.get("video", "")
    if video_url:
        print(f"   Descargando video...")
        video_ok = descargar_media(
            video_url, video_folder, "video_producto.mp4",
            referer_url=url_producto, es_video=True,
        )
        print(f"   Video: {'✓' if video_ok else '✗ falló'}")
    else:
        print("   Sin video en este producto")

    # Guardar specs en txt
    specs = data.get("specs", [])
    if specs:
        with open(os.path.join(product_path, "especificaciones.txt"), "w", encoding="utf-8") as f:
            f.write("\n".join(specs))

    # Calcular precios
    try:
        precio_cny = float(data.get("precio") or 0)
    except Exception:
        precio_cny = 0.0
    precio_usd   = round(precio_cny / tasa, 2) if precio_cny else ""
    precio_venta = round(precio_usd * MARGEN, 2) if precio_usd else ""

    # Limpiar nombre de tienda
    tienda_raw   = data.get("tienda", "")
    tienda_limpia = re.sub(r'\d+\.\d+', '', tienda_raw)
    tienda_limpia = re.sub(r'好评率\d+%|平均\d+小时发货|客服满意度\d+%|88VIP', '', tienda_limpia).strip()

    nombre_zh = data.get("nombre", "") or ""
    nombre_en = nombre_zh
    if nombre_zh and any("\u4e00" <= c <= "\u9fff" for c in nombre_zh):
        print("   Traduciendo nombre ZH→EN...")
        tr = traducir_zh_a_en(nombre_zh)
        nombre_en = tr if tr else "(traducir)"
    elif not nombre_zh:
        nombre_en = ""

    tienda_rec = data.get("tienda_recomendados") or []
    tienda_rec_json = json.dumps(tienda_rec, ensure_ascii=False)[:2000] if tienda_rec else ""

    # Construir fila Excel
    nueva_fila = {
        "Tienda":           tienda_limpia,
        "Nombre (ZH)":      nombre_zh,
        "Nombre (EN)":      nombre_en,
        "Descripcion":      (data.get("descripcion") or "")[:400],
        "Specs":            " | ".join(specs)[:400],
        "Variaciones":      " | ".join(data.get("variaciones", []))[:300],
        "Costo CNY":        precio_cny or "",
        "Costo USD":        precio_usd,
        "Precio Venta USD": precio_venta,
        "Calificacion":     data.get("calificaciones", ""),
        "Sitio":            data.get("sitio", ""),
        "URL":              url_producto,
        "Imagen1": imagenes_urls[0] if len(imagenes_urls) > 0 else "",
        "Imagen2": imagenes_urls[1] if len(imagenes_urls) > 1 else "",
        "Imagen3": imagenes_urls[2] if len(imagenes_urls) > 2 else "",
        "Imagen4": imagenes_urls[3] if len(imagenes_urls) > 3 else "",
        "Imagen5": imagenes_urls[4] if len(imagenes_urls) > 4 else "",
        "Imagen6": imagenes_urls[5] if len(imagenes_urls) > 5 else "",
        "Imagen7": imagenes_urls[6] if len(imagenes_urls) > 6 else "",
        "Imagen8": imagenes_urls[7] if len(imagenes_urls) > 7 else "",
        "Fotos Descargadas": imagenes_ok,
        "Variantes URLs": " | ".join((data.get("imagenes_variantes") or [])[:12]),
        "Variantes Descargadas": var_ok if imagenes_variantes else "",
        "本店推荐 JSON": tienda_rec_json,
        "Img Desc1": desc_urls[0] if len(desc_urls) > 0 else "",
        "Img Desc2": desc_urls[1] if len(desc_urls) > 1 else "",
        "Img Desc3": desc_urls[2] if len(desc_urls) > 2 else "",
        "Imgs Desc Descargadas": desc_ok,
        "Video URL":      video_url,
        "Video Descargado": "Sí" if video_ok else "",
        "Carpeta Local":  folder_name,
        "Tasa CNY/USD":   round(tasa, 4),
        "Fecha":          datetime.now().strftime("%Y-%m-%d %H:%M"),
    }

    for k, v in (data.get("datos_custom") or {}).items():
        nueva_fila[f"CUSTOM_{k}"] = (" | ".join(v) if isinstance(v, list) else str(v or ""))[:200]

    hacer_backup_excel()

    df_nuevo = pd.DataFrame([nueva_fila])
    if df_exist is not None and not df_exist.empty:
        for col in df_nuevo.columns:
            if col not in df_exist.columns:
                df_exist[col] = ""
        for col in df_exist.columns:
            if col not in df_nuevo.columns:
                df_nuevo[col] = ""
        df_final = pd.concat([df_exist, df_nuevo], ignore_index=True)
    else:
        df_final = df_nuevo

    try:
        df_final.to_excel(EXCEL_PATH, index=False, engine="openpyxl")
        print("   Excel maestro actualizado ✓")
    except PermissionError:
        msg = "❌ Excel maestro está abierto. Ciérralo y reintenta."
        print(f"   {msg}")
        return jsonify({"status": "error", "message": msg}), 409

    aplicar_formato_excel(EXCEL_PATH)

    total_filas = len(df_final)
    print(f"   Total en maestro: {total_filas} productos")
    print("=" * 60)

    return jsonify({
        "status":        "success",
        "message":       (
            f"✅ Guardado | "
            f"{imagenes_ok}/{len(imagenes_urls)} fotos | "
            f"{desc_ok}/{len(desc_urls)} imgs desc | "
            f"{'Video ✓' if video_ok else 'Sin video'} | "
            f"Total: {total_filas} productos"
        ),
        "carpeta":       folder_name,
        "total_maestro": total_filas,
        "tasa_usada":    round(tasa, 4),
    })


# ─── STATS ───────────────────────────────────────────────────
@app.route('/stats', methods=['GET'])
def ver_stats():
    if not os.path.exists(EXCEL_PATH):
        return jsonify({"total": 0, "mensaje": "Excel maestro aún no creado"})
    try:
        df = pd.read_excel(EXCEL_PATH, engine="openpyxl")
        return jsonify({
            "total":    len(df),
            "columnas": list(df.columns),
            "ultimo":   df.iloc[-1].to_dict() if not df.empty else {},
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── INICIO ──────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("  BIO CATTALEYA SCRAPER PRO — RECEPTOR v4.1")
    print("=" * 60)
    print(f"  Carpeta raíz : {BASE_DIR}")
    print(f"  Output       : {OUTPUT_DIR}")
    print(f"  Excel maestro: {EXCEL_PATH}")
    print(f"  Backups      : {BACKUP_DIR}")
    print(f"  Puerto       : http://127.0.0.1:5000")
    print("=" * 60)
    print("  Estructura de cada producto:")
    print("  output / [fecha_nombre] /")
    print("                imagenes /      ← fotos del producto")
    print("                imagenes_variantes / ← colores/SKU")
    print("                videos /        ← video si existe")
    print("                imagenes_descripcion /")
    print("                datos_completo.json")
    print("                especificaciones.txt")
    print("=" * 60)
    app.run(port=5000, debug=False)
